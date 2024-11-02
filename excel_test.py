# import openpyxl
# from openpyxl.styles import Font, Alignment
# from datetime import datetime
# import re

# # Загружаем книги
# wb = openpyxl.load_workbook('/Users/seu/python/excel/Реестр договоров_оригинал_28.10.xlsx')
# wb2 = openpyxl.load_workbook('/Users/seu/python/excel/Реестр_Final (Копия для Документарного отдела).xlsm')

# # Выбираем листы
# wb_sheet = wb['Реестр']
# wb_sheet2 = wb2['СВОДНАЯ']

# # Столбцы для копирования (индексы)
# copy_columns = [2, 1, 8, 9, 14]  # Пример

# # Функция для нахождения первой пустой строки
# def find_first_empty_row(sheet):
#     for row in range(1, sheet.max_row + 2):  # +2, чтобы учесть ситуацию с добавлением новой строки
#         if sheet.cell(row=row, column=1).value is None:  # Проверяем только первый столбец
#             return row
#     return sheet.max_row + 1  # Если все строки заполнены, возвращаем следующую строку

# # Считываем данные
# data = []
# beneficiary_dates = {} 

# # Основной алгоритм
# for row in wb_sheet2.iter_rows(min_row=2, values_only=True):
#     selected_data = [row[i] for i in copy_columns]

#     if selected_data[4]:  # Проверяем, что ячейка не пустая
#         date_str = None  # Сбросим дату для каждой строки

#         # Проверяем каждый паттерн
# patterns = [
    # r'(\d{1,2})(\d{2})(\d{2})',          # ddMMyy
    # r'(\d{1,2})\.(\d{1,2})\.(\d{2})',    # dd.mm.yy
    # r'(\d{1,2})/(\d{1,2})/(\d{4})',      # dd/mm/yyyy
    # r'(\d{1,2})-(\d{1,2})-(\d{4})',      # dd-mm-yyyy
    # r'(\d{1,2})\s+(\w+)\s+(\d{4})',      # dd Month yyyy
    # r'(\w+)\s+(\d{1,2}),\s+(\d{4})',     # Month dd, yyyy
    # r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})',   # dd.mm.yy или dd.mm.yyyy
    # r'(\d{1,2})/(\d{1,2})\.(\d{2,4})',    # dd/mm.yy или dd/mm.yyyy
    # r'(\d{1,2})\s+(\d{1,2})\s*,\s*(\d{4})', # dd, dd Month yyyy
    # r'dd\s+(\d{1,2})\.(\d{1,2})\.(\d{4})',  # dd dd.mm.yyyy
    # r'dated\s+(\d{1,2})\s+(\w+)\s+(\d{4})', # dated dd Month yyyy
    # r'dd\s+(\d{1,2})\s+(\w+)\s+(\d{4})',    # dd Month yyyy
    # r'dd.\s(\d{1,2})(\d{2})(\d{2})',          # dd. ddMMyy
    # r'dd.\s(\d{1,2})\.(\d{1,2})\.(\d{2})',    # dd. dd.mm.yy
    # r'dd.\s(\d{1,2})/(\d{1,2})/(\d{4})',      # dd. dd/mm/yyyy
    # r'dd.\s(\d{1,2})-(\d{1,2})-(\d{4})',      # dd. dd-mm-yyyy
    # r'dd.\s(\d{1,2})\s+(\w+)\s+(\d{4})',      # dd. dd Month yyyy
    # r'dd.\s(\w+)\s+(\d{1,2}),\s+(\d{4})',     # dd. Month dd, yyyy
    # r'dd.\s(\d{1,2})\.(\d{1,2})\.(\d{2,4})',   # dd. dd.mm.yy или dd.mm.yyyy
    # r'dd.\s(\d{1,2})/(\d{1,2})\.(\d{2,4})',    # dd. dd/mm.yy или dd/mm.yyyy
    # r'dated\s+(\d{1,2})\s+(\w+)\s+(\d{4})',    # dated Month yyyy
    # r'dated\s(\d{1,2})(\d{2})(\d{2})',          # dated ddMMyy
    # r'dated\s(\d{1,2})\.(\d{1,2})\.(\d{2})',    # dated dd.mm.yy
    # r'dated\s(\d{1,2})/(\d{1,2})/(\d{4})',      # dated dd/mm/yyyy
    # r'dated\s(\d{1,2})-(\d{1,2})-(\d{4})',      # dated dd-mm-yyyy
    # r'dated\s(\d{1,2})\s+(\w+)\s+(\d{4})',      # dated dd Month yyyy
    # r'dated\s(\w+)\s+(\d{1,2}),\s+(\d{4})',     # dated Month dd, yyyy
    # r'dated\s(\d{1,2})\.(\d{1,2})\.(\d{2,4})',   # dated dd.mm.yy или dd.mm.yyyy
    # r'dated\s(\d{1,2})/(\d{1,2})\.(\d{2,4})',    # dated dd/mm.yy или dd/mm.yyyy
    # r'dd\s*(\d{1,2})\.(\d{1,2})\.(\d{2,4})',     # dd dd.mm.yyyy
    # r'(\d{1,2})/(\d{1,2})/(\d{2})',              # dd/mm/yy
    # r'(\d{1,2})/(\d{1,2})/(\d{2,4})',            # dd/mm/yyyy
    # r'(\d{1,2})\s*-\s*(\d{1,2})\s*-\s*(\d{4})',
    # r'(?:dd\s+)?(\d{1,2})\.(\d{1,2})\.(\d{2}|\d{4})',  # dd.mm.yy или dd.mm.yyyy
    # r'(?:dd\s+)?(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})',      # dd/mm.yy или dd/mm.yyyy
    # r'(?:dd\s+)?(\d{1,2})-(\d{1,2})-(\d{2}|\d{4})',      # dd-mm.yy или dd-mm.yyyy   # dd-mm-yyyy 
    # r'(\d{4})/(\d{1,2})/(\d{1,2})',                 # YYYY/MM/dd
    # r'(\d{4})-(\d{1,2})-(\d{1,2})',                 # YYYY-MM-dd
    # r'(\d{1,2})\s+(\w+)\s+(\d{4})',                 # dd Month yyyy (с пробелами)
    # r'(?:(?:dd|DD|dd\.)?\s*(\d{1,2})\.(\d{1,2})\.(\d{2}|\d{4}))',  # dd.mm.yy или dd.mm.yyyy с dd/DD
    # r'(?:(?:dd|DD|dd\.)?\s*(\d{1,2})/(\d{1,2})/(\d{2}|\d{4}))',    # dd/mm.yy или dd/mm.yyyy с dd/DD
    # r'(?:(?:dd|DD|dd\.)?\s*(\d{1,2})-(\d{1,2})-(\d{2}|\d{4}))',    # dd-mm.yy или dd-mm.yyyy с dd/DD
    # r'(?:\bdd\.?\s*)?(\d{1,2})\s*(\w+)\s*(\d{4})',                # dd Month yyyy с пробелами
    # r'(?:\bdd\.?\s*)?(\w+)\s+(\d{1,2}),\s+(\d{4})',              # Month dd, yyyy
# ]

# # Обработка строк из Excel
# for row in wb_sheet2.iter_rows(min_row=2, values_only=True):
#     selected_data = [row[i] for i in copy_columns]

#     if selected_data[4]:  # Проверяем, что ячейка не пустая
#         date_str = None  # Сбросим дату для каждой строки

#         # Проверяем каждый паттерн
#         for pattern in patterns:
#             match = re.search(pattern, selected_data[4], re.IGNORECASE)
#             if match:
#                 if len(match.groups()) == 3:
#                     if pattern == patterns[0]:  # ddMMyy
#                         date_str = f"{match.group(1)}.{match.group(2)}.{match.group(3)}"
#                     elif pattern in [patterns[1], patterns[3]]:  # dd.mm.yy или dd-mm-yyyy
#                         date_str = f"{match.group(1)}.{match.group(2)}.{int(match.group(3)) + 2000}"
#                     elif pattern == patterns[2]:  # dd/mm/yyyy
#                         date_str = selected_data[4]  # Используем исходную строку
#                     elif pattern == patterns[4]:  # dd Month yyyy
#                         month_map = {
#                             "January": "01", "February": "02", "March": "03",
#                             "April": "04", "May": "05", "June": "06",
#                             "July": "07", "August": "08",
#                             "September": "09", "October": "10",
#                             "November": "11", "December": "12"
#                         }
#                         month = month_map.get(match.group(2), "00")
#                         date_str = f"{match.group(1)}.{month}.{match.group(3)}"
#                     elif pattern == patterns[5]:  # Month dd, yyyy
#                         month_map = {
#                             "January": "01", "February": "02", "March": "03",
#                             "April": "04", "May": "05", "June": "06",
#                             "July": "07", "August": "08",
#                             "September": "09", "October": "10",
#                             "November": "11", "December": "12"
#                         }
#                         month = month_map.get(match.group(1), "00")
#                         date_str = f"{match.group(2)}.{month}.{match.group(3)}"
#                     elif pattern in [patterns[6], patterns[7]]:
#                         date_str = f"{match.group(1)}.{match.group(2)}.{match.group(3)}"
#                     elif pattern == patterns[8]:
#                         date_str = f"{match.group(1)}.{match.group(2)}.{match.group(3)}"
#                 elif len(match.groups()) == 2:
#                     if pattern in [patterns[9], patterns[10], patterns[11]]:
#                         month_map = {
#                             "January": "01", "February": "02", "March": "03",
#                             "April": "04", "May": "05", "June": "06",
#                             "July": "07", "August": "08",
#                             "September": "09", "October": "10",
#                             "November": "11", "December": "12"
#                         }
#                         month = month_map.get(match.group(2), "00")
#                         date_str = f"{match.group(1)}.{month}.{match.group(3)}"
#                 break  # Если нашли дату, выходим из цикла

#         # Преобразуем найденную строку в дату
#         if date_str:
#             try:
#                 date_value = datetime.strptime(date_str, '%d.%m.%Y').date()
#                 selected_data[4] = date_value
#             except ValueError:
#                 print(f"Неверный формат даты: {date_str}")

#     data.append(selected_data)

# print("Данные для записи:", data)


# # Находим первую пустую строку
# start_row = find_first_empty_row(wb_sheet)
# print(f"Начальная строка для записи: {start_row}")

# # Записываем данные в целевой лист
# for i, row in enumerate(data):
#     for j, value in enumerate(row):
#         cell = wb_sheet.cell(row=start_row + i, column=j + 1, value=value)

#         # Устанавливаем жирный шрифт
#         cell.font = Font(bold=True)

#         # Центрируем текст
#         cell.alignment = Alignment(horizontal='center')

# # Устанавливаем ширину столбцов автоматически
# for column in wb_sheet.columns:
#     max_length = 0
#     column = [cell for cell in column]
#     for cell in column:
#         try:
#             if len(str(cell.value)) > max_length:
#                 max_length = len(str(cell.value))
#         except:
#             pass
#     adjusted_width = (max_length + 2)
#     wb_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

# # Сохраняем изменения
# wb.save('/Users/seu/python/excel/Реестр договоров_оригинал_28.10.xlsx')
# print("Данные успешно записаны.")


import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
import re

# Загружаем книги
wb = openpyxl.load_workbook('/Users/seu/python/excel/Реестр договоров_оригинал_28.10.xlsx')
wb2 = openpyxl.load_workbook('/Users/seu/python/excel/Реестр_Final (Копия для Документарного отдела).xlsm')

# Выбираем листы
wb_sheet = wb['Реестр']
wb_sheet2 = wb2['СВОДНАЯ']

# Столбцы для копирования (индексы)
copy_columns = [2, 1, 8, 9, 14]  # Пример

# Функция для нахождения первой пустой строки
def find_first_empty_row(sheet):
    for row in range(1, sheet.max_row + 2):  # +2, чтобы учесть ситуацию с добавлением новой строки
        if sheet.cell(row=row, column=1).value is None:  # Проверяем только первый столбец
            return row
    return sheet.max_row + 1  # Если все строки заполнены, возвращаем следующую строку

# Считываем данные
data = []
beneficiary_dates = {}

# Шаблоны для поиска дат
patterns = [
    r'(\d{1,2})(\d{2})(\d{2})',          # ddMMyy
    r'(\d{1,2})\.(\d{1,2})\.(\d{2})',    # dd.mm.yy
    r'(\d{1,2})/(\d{1,2})/(\d{4})',      # dd/mm/yyyy
    r'(\d{1,2})-(\d{1,2})-(\d{4})',      # dd-mm-yyyy
    r'(\d{1,2})\s+(\w+)\s+(\d{4})',      # dd Month yyyy
    r'(\w+)\s+(\d{1,2}),\s+(\d{4})',     # Month dd, yyyy
    r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})',   # dd.mm.yy или dd.mm.yyyy
    r'(\d{1,2})/(\d{1,2})\.(\d{2,4})',    # dd/mm.yy или dd/mm.yyyy
    r'(\d{1,2})\s+(\d{1,2})\s*,\s*(\d{4})', # dd, dd Month yyyy
    r'dd\s+(\d{1,2})\.(\d{1,2})\.(\d{4})',  # dd dd.mm.yyyy
    r'dated\s+(\d{1,2})\s+(\w+)\s+(\d{4})', # dated dd Month yyyy
    r'dd\s+(\d{1,2})\s+(\w+)\s+(\d{4})',    # dd Month yyyy
    r'dd.\s(\d{1,2})(\d{2})(\d{2})',          # dd. ddMMyy
    r'dd.\s(\d{1,2})\.(\d{1,2})\.(\d{2})',    # dd. dd.mm.yy
    r'dd.\s(\d{1,2})/(\d{1,2})/(\d{4})',      # dd. dd/mm/yyyy
    r'dd.\s(\d{1,2})-(\d{1,2})-(\d{4})',      # dd. dd-mm-yyyy
    r'dd.\s(\d{1,2})\s+(\w+)\s+(\d{4})',      # dd. dd Month yyyy
    r'dd.\s(\w+)\s+(\d{1,2}),\s+(\d{4})',     # dd. Month dd, yyyy
    r'dd.\s(\d{1,2})\.(\d{1,2})\.(\d{2,4})',   # dd. dd.mm.yy или dd.mm.yyyy
    r'dd.\s(\d{1,2})/(\d{1,2})\.(\d{2,4})',    # dd. dd/mm.yy или dd/mm.yyyy
    r'dated\s+(\d{1,2})\s+(\w+)\s+(\d{4})',    # dated Month yyyy
    r'dated\s(\d{1,2})(\d{2})(\d{2})',          # dated ddMMyy
    r'dated\s(\d{1,2})\.(\d{1,2})\.(\d{2})',    # dated dd.mm.yy
    r'dated\s(\d{1,2})/(\d{1,2})/(\d{4})',      # dated dd/mm/yyyy
    r'dated\s(\d{1,2})-(\d{1,2})-(\d{4})',      # dated dd-mm-yyyy
    r'dated\s(\d{1,2})\s+(\w+)\s+(\d{4})',      # dated dd Month yyyy
    r'dated\s(\w+)\s+(\d{1,2}),\s+(\d{4})',     # dated Month dd, yyyy
    r'dated\s(\d{1,2})\.(\d{1,2})\.(\d{2,4})',   # dated dd.mm.yy или dd.mm.yyyy
    r'dated\s(\d{1,2})/(\d{1,2})\.(\d{2,4})',    # dated dd/mm.yy или dd/mm.yyyy
    r'dd\s*(\d{1,2})\.(\d{1,2})\.(\d{2,4})',     # dd dd.mm.yyyy
    r'(\d{1,2})/(\d{1,2})/(\d{2})',              # dd/mm/yy
    r'(\d{1,2})/(\d{1,2})/(\d{2,4})',            # dd/mm/yyyy
    r'(\d{1,2})\s*-\s*(\d{1,2})\s*-\s*(\d{4})',
    r'(?:dd\s+)?(\d{1,2})\.(\d{1,2})\.(\d{2}|\d{4})',  # dd.mm.yy или dd.mm.yyyy
    r'(?:dd\s+)?(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})',      # dd/mm.yy или dd/mm.yyyy
    r'(?:dd\s+)?(\d{1,2})-(\d{1,2})-(\d{2}|\d{4})',      # dd-mm.yy или dd-mm.yyyy   # dd-mm-yyyy 
    r'(\d{4})/(\d{1,2})/(\d{1,2})',                 # YYYY/MM/dd
    r'(\d{4})-(\d{1,2})-(\d{1,2})',                 # YYYY-MM-dd
    r'(\d{1,2})\s+(\w+)\s+(\d{4})',                 # dd Month yyyy (с пробелами)
    r'(?:(?:dd|DD|dd\.)?\s*(\d{1,2})\.(\d{1,2})\.(\d{2}|\d{4}))',  # dd.mm.yy или dd.mm.yyyy с dd/DD
    r'(?:(?:dd|DD|dd\.)?\s*(\d{1,2})/(\d{1,2})/(\d{2}|\d{4}))',    # dd/mm.yy или dd/mm.yyyy с dd/DD
    r'(?:(?:dd|DD|dd\.)?\s*(\d{1,2})-(\d{1,2})-(\d{2}|\d{4}))',    # dd-mm.yy или dd-mm.yyyy с dd/DD
    r'(?:\bdd\.?\s*)?(\d{1,2})\s*(\w+)\s*(\d{4})',                # dd Month yyyy с пробелами
    r'(?:\bdd\.?\s*)?(\w+)\s+(\d{1,2}),\s+(\d{4})',              # Month dd, yyyy
]

# Обработка строк из Excel
for row in wb_sheet2.iter_rows(min_row=2, values_only=True):
    selected_data = [row[i] for i in copy_columns]

    if selected_data[4]:  # Проверяем, что ячейка с датой не пустая
        date_str = None  # Сбросим дату для каждой строки

        # Проверяем каждый паттерн
        for pattern in patterns:
            match = re.search(pattern, selected_data[4], re.IGNORECASE)
            if match:
                if len(match.groups()) == 3:
                    # Обработка разных форматов
                    if pattern == patterns[0]:  # ddMMyy
                        date_str = f"{match.group(1)}.{match.group(2)}.{match.group(3)}"
                    elif pattern in [patterns[1], patterns[3]]:  # dd.mm.yy или dd-mm-yyyy
                        date_str = f"{match.group(1)}.{match.group(2)}.{int(match.group(3)) + 2000}"
                    elif pattern == patterns[2]:  # dd/mm/yyyy
                        date_str = selected_data[4]  # Используем исходную строку
                    elif pattern == patterns[4]:  # dd Month yyyy
                        month_map = {
                            "January": "01", "February": "02", "March": "03",
                            "April": "04", "May": "05", "June": "06",
                            "July": "07", "August": "08",
                            "September": "09", "October": "10",
                            "November": "11", "December": "12"
                        }
                        month = month_map.get(match.group(2), "00")
                        date_str = f"{match.group(1)}.{month}.{match.group(3)}"
                    elif pattern == patterns[5]:  # Month dd, yyyy
                        month_map = {
                            "January": "01", "February": "02", "March": "03",
                            "April": "04", "May": "05", "June": "06",
                            "July": "07", "August": "08",
                            "September": "09", "October": "10",
                            "November": "11", "December": "12"
                        }
                        month = month_map.get(match.group(1), "00")
                        date_str = f"{match.group(2)}.{month}.{match.group(3)}"
                break  # Если нашли дату, выходим из цикла

        # Преобразуем найденную строку в дату
        if date_str:
            try:
                date_value = datetime.strptime(date_str, '%d.%m.%Y').date()
                beneficiary = selected_data[3]  # Бенефициар в 9-м столбце (индекс 3)

                # Сохраняем только самую раннюю дату для каждого бенефициара
                if beneficiary not in beneficiary_dates or date_value < beneficiary_dates[beneficiary]:
                    beneficiary_dates[beneficiary] = date_value
                    selected_data[4] = date_value  # Заменяем дату в выбранных данных
                    data.append(selected_data)  # Добавляем всю строку

            except ValueError:
                print(f"Неверный формат даты: {date_str}")

# Находим первую пустую строку
start_row = find_first_empty_row(wb_sheet)
print(f"Начальная строка для записи: {start_row}")

# Записываем данные в целевой лист
for i, row in enumerate(data):
    for j, value in enumerate(row):
        cell = wb_sheet.cell(row=start_row + i, column=j + 1, value=value)

        # Устанавливаем жирный шрифт
        cell.font = Font(bold=True)

        # Центрируем текст
        cell.alignment = Alignment(horizontal='center')

# Устанавливаем ширину столбцов автоматически
for column in wb_sheet.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    wb_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

# Сохраняем изменения
wb.save('/Users/seu/python/excel/Реестр договоров_оригинал_28.10.xlsx')
print("Данные успешно записаны.")
