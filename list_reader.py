import sys  # sys нужен для передачи argv в QApplication
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QPushButton, QMessageBox
import excel_reader
import os
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
import re

	
class ExampleApp(QtWidgets.QMainWindow, excel_reader.Ui_MainWindow):
    def __init__(self):

        self.folder_give = []
        self.folder_take = []
        # Это здесь нужно для доступа к переменным, методам
        # и т.д. в файле design.py
        super().__init__()
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна

        self.toolButton.clicked.connect(self.first_folder)
        self.toolButton_2.clicked.connect(self.second_folder)
        self.pushButton.clicked.connect(self.take_benificuarus)

    def first_folder(self):
        self.textBrowser.clear()
        file_name = QtWidgets.QFileDialog.getOpenFileName(self, "Выберите файл")
        print(file_name)
        self.folder_give.append(file_name[0])

        if file_name:  # Проверяем, был ли выбран файл
            self.textBrowser.append(file_name[0]) 

    def second_folder(self):
        self.textBrowser_2.clear()
        file_name2 = QtWidgets.QFileDialog.getOpenFileName(self, "Выберите файл")
        print(file_name2)
        self.folder_take.append(file_name2[0])

        if file_name2:  # Проверяем, был ли выбран файл
            self.textBrowser_2.append(file_name2[0]) 

    def take_benificuarus(self):

        # Загружаем книги
        # wb = openpyxl.load_workbook('/Users/seu/python/excel/Реестр договоров_оригинал_28.10.xlsx')
        # wb2 = openpyxl.load_workbook('/Users/seu/python/excel/Реестр_Final (Копия для Документарного отдела).xlsm')
        wb = openpyxl.load_workbook(self.folder_take[0])
        wb2= openpyxl.load_workbook(self.folder_give[0])

        # Выбираем листы
        wb_sheet = wb['Реестр']
        wb_sheet2 = wb2['СВОДНАЯ']

        # Столбцы для копирования (индексы)
        copy_columns = [2, 1, 8, 9, 14]  # Пример

        # Функция для нахождения первой пустой строки
        def find_first_empty_row(sheet):
            for row in range(1, sheet.max_row + 2):  # +2, чтобы учесть ситуацию с добавлением новой строки
                if sheet.cell(row=row, column=1).value is None:  # Проверяем только первый столбец
                    return row
            return sheet.max_row + 1  # Если все строки заполнены, возвращаем следующую строку

        # Считываем данные
        data = []
        beneficiary_dates = {}

        # Шаблоны для поиска дат
        patterns = [
            r'(\d{1,2})(\d{2})(\d{2})',          # ddMMyy
            r'(\d{1,2})\.(\d{1,2})\.(\d{2})',    # dd.mm.yy
            r'(\d{1,2})/(\d{1,2})/(\d{4})',      # dd/mm/yyyy
            r'(\d{1,2})-(\d{1,2})-(\d{4})',      # dd-mm-yyyy
            r'(\d{1,2})\s+(\w+)\s+(\d{4})',      # dd Month yyyy
            r'(\w+)\s+(\d{1,2}),\s+(\d{4})',     # Month dd, yyyy
            r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})',   # dd.mm.yy или dd.mm.yyyy
            r'(\d{1,2})/(\d{1,2})\.(\d{2,4})',    # dd/mm.yy или dd/mm.yyyy
            r'(\d{1,2})\s+(\d{1,2})\s*,\s*(\d{4})', # dd, dd Month yyyy
            r'dd\s+(\d{1,2})\.(\d{1,2})\.(\d{4})',  # dd dd.mm.yyyy
            r'dated\s+(\d{1,2})\s+(\w+)\s+(\d{4})', # dated dd Month yyyy
            r'dd\s+(\d{1,2})\s+(\w+)\s+(\d{4})',    # dd Month yyyy
            r'dd.\s(\d{1,2})(\d{2})(\d{2})',          # dd. ddMMyy
            r'dd.\s(\d{1,2})\.(\d{1,2})\.(\d{2})',    # dd. dd.mm.yy
            r'dd.\s(\d{1,2})/(\d{1,2})/(\d{4})',      # dd. dd/mm/yyyy
            r'dd.\s(\d{1,2})-(\d{1,2})-(\d{4})',      # dd. dd-mm-yyyy
            r'dd.\s(\d{1,2})\s+(\w+)\s+(\d{4})',      # dd. dd Month yyyy
            r'dd.\s(\w+)\s+(\d{1,2}),\s+(\d{4})',     # dd. Month dd, yyyy
            r'dd.\s(\d{1,2})\.(\d{1,2})\.(\d{2,4})',   # dd. dd.mm.yy или dd.mm.yyyy
            r'dd.\s(\d{1,2})/(\d{1,2})\.(\d{2,4})',    # dd. dd/mm.yy или dd/mm.yyyy
            r'dated\s+(\d{1,2})\s+(\w+)\s+(\d{4})',    # dated Month yyyy
            r'dated\s(\d{1,2})(\d{2})(\d{2})',          # dated ddMMyy
            r'dated\s(\d{1,2})\.(\d{1,2})\.(\d{2})',    # dated dd.mm.yy
            r'dated\s(\d{1,2})/(\d{1,2})/(\d{4})',      # dated dd/mm/yyyy
            r'dated\s(\d{1,2})-(\d{1,2})-(\d{4})',      # dated dd-mm-yyyy
            r'dated\s(\d{1,2})\s+(\w+)\s+(\d{4})',      # dated dd Month yyyy
            r'dated\s(\w+)\s+(\d{1,2}),\s+(\d{4})',     # dated Month dd, yyyy
            r'dated\s(\d{1,2})\.(\d{1,2})\.(\d{2,4})',   # dated dd.mm.yy или dd.mm.yyyy
            r'dated\s(\d{1,2})/(\d{1,2})\.(\d{2,4})',    # dated dd/mm.yy или dd/mm.yyyy
            r'dd\s*(\d{1,2})\.(\d{1,2})\.(\d{2,4})',     # dd dd.mm.yyyy
            r'(\d{1,2})/(\d{1,2})/(\d{2})',              # dd/mm/yy
            r'(\d{1,2})/(\d{1,2})/(\d{2,4})',            # dd/mm/yyyy
            r'(\d{1,2})\s*-\s*(\d{1,2})\s*-\s*(\d{4})',
            r'(?:dd\s+)?(\d{1,2})\.(\d{1,2})\.(\d{2}|\d{4})',  # dd.mm.yy или dd.mm.yyyy
            r'(?:dd\s+)?(\d{1,2})/(\d{1,2})/(\d{2}|\d{4})',      # dd/mm.yy или dd/mm.yyyy
            r'(?:dd\s+)?(\d{1,2})-(\d{1,2})-(\d{2}|\d{4})',      # dd-mm.yy или dd-mm.yyyy   # dd-mm-yyyy 
            r'(\d{4})/(\d{1,2})/(\d{1,2})',                 # YYYY/MM/dd
            r'(\d{4})-(\d{1,2})-(\d{1,2})',                 # YYYY-MM-dd
            r'(\d{1,2})\s+(\w+)\s+(\d{4})',                 # dd Month yyyy (с пробелами)
            r'(?:(?:dd|DD|dd\.)?\s*(\d{1,2})\.(\d{1,2})\.(\d{2}|\d{4}))',  # dd.mm.yy или dd.mm.yyyy с dd/DD
            r'(?:(?:dd|DD|dd\.)?\s*(\d{1,2})/(\d{1,2})/(\d{2}|\d{4}))',    # dd/mm.yy или dd/mm.yyyy с dd/DD
            r'(?:(?:dd|DD|dd\.)?\s*(\d{1,2})-(\d{1,2})-(\d{2}|\d{4}))',    # dd-mm.yy или dd-mm.yyyy с dd/DD
            r'(?:\bdd\.?\s*)?(\d{1,2})\s*(\w+)\s*(\d{4})',                # dd Month yyyy с пробелами
            r'(?:\bdd\.?\s*)?(\w+)\s+(\d{1,2}),\s+(\d{4})',              # Month dd, yyyy
        ]

        # Обработка строк из Excel
        for row in wb_sheet2.iter_rows(min_row=2, values_only=True):
            selected_data = [row[i] for i in copy_columns]

            if selected_data[4]:  # Проверяем, что ячейка с датой не пустая
                date_str = None  # Сбросим дату для каждой строки

                # Проверяем каждый паттерн
                for pattern in patterns:
                    match = re.search(pattern, selected_data[4], re.IGNORECASE)
                    if match:
                        if len(match.groups()) == 3:
                            # Обработка разных форматов
                            if pattern == patterns[0]:  # ddMMyy
                                date_str = f"{match.group(1)}.{match.group(2)}.{match.group(3)}"
                            elif pattern in [patterns[1], patterns[3]]:  # dd.mm.yy или dd-mm-yyyy
                                date_str = f"{match.group(1)}.{match.group(2)}.{int(match.group(3)) + 2000}"
                            elif pattern == patterns[2]:  # dd/mm/yyyy
                                date_str = selected_data[4]  # Используем исходную строку
                            elif pattern == patterns[4]:  # dd Month yyyy
                                month_map = {
                                    "January": "01", "February": "02", "March": "03",
                                    "April": "04", "May": "05", "June": "06",
                                    "July": "07", "August": "08",
                                    "September": "09", "October": "10",
                                    "November": "11", "December": "12"
                                }
                                month = month_map.get(match.group(2), "00")
                                date_str = f"{match.group(1)}.{month}.{match.group(3)}"
                            elif pattern == patterns[5]:  # Month dd, yyyy
                                month_map = {
                                    "January": "01", "February": "02", "March": "03",
                                    "April": "04", "May": "05", "June": "06",
                                    "July": "07", "August": "08",
                                    "September": "09", "October": "10",
                                    "November": "11", "December": "12"
                                }
                                month = month_map.get(match.group(1), "00")
                                date_str = f"{match.group(2)}.{month}.{match.group(3)}"
                        break  # Если нашли дату, выходим из цикла

                # Преобразуем найденную строку в дату
                if date_str:
                    try:
                        date_value = datetime.strptime(date_str, '%d.%m.%Y').date()
                        beneficiary = selected_data[3]  # Бенефициар в 9-м столбце (индекс 3)

                        # Сохраняем только самую раннюю дату для каждого бенефициара
                        if beneficiary not in beneficiary_dates or date_value < beneficiary_dates[beneficiary]:
                            beneficiary_dates[beneficiary] = date_value
                            selected_data[4] = date_value  # Заменяем дату в выбранных данных
                            data.append(selected_data)  # Добавляем всю строку

                    except ValueError:
                        print(f"Неверный формат даты: {date_str}")

        # Находим первую пустую строку
        start_row = find_first_empty_row(wb_sheet)
        print(f"Начальная строка для записи: {start_row}")

        # Записываем данные в целевой лист
        for i, row in enumerate(data):
            for j, value in enumerate(row):
                cell = wb_sheet.cell(row=start_row + i, column=j + 1, value=value)

                # Устанавливаем жирный шрифт
                cell.font = Font(bold=True)

                # Центрируем текст
                cell.alignment = Alignment(horizontal='center')

        # Устанавливаем ширину столбцов автоматически
        for column in wb_sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Сохраняем изменения
        wb.save('/Users/seu/python/excel/Реестр договоров_оригинал_28.10.xlsx')
        print("Данные успешно записаны.")
        return self.showQmsg()

    def showQmsg(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText("Бенифициары успешно определены!")
        msg.setWindowTitle("Информация")
        msg.exec_()
        # msg.buttonClicked.connect()
        # returnValue = msg.exec_()
        # if returnValue == QMessageBox.Ok:
        #     print('OK clicked')


	
def main():
    app = QtWidgets.QApplication(sys.argv)  # Новый экземпляр QApplication
    window = ExampleApp()  # Создаём объект класса ExampleApp
    window.show()  # Показываем окно
    app.exec_()  # и запускаем приложение

if __name__ == '__main__':  # Если мы запускаем файл напрямую, а не импортируем
    main()  # то запускаем функцию main()
